import datetime
import smtplib
import ssl
from email.message import EmailMessage
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from email.mime.text import MIMEText

import os
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


def sendMail():
    email_sender = 'philippsiebenforcher@gmail.com'
    email_password = ''
    email_receiver = ''

    subject = "Brotbestellung für den " + str(datetime.date.today())
    body = """
    """

    # Create a multipart email message
    em = MIMEMultipart()
    em['from'] = email_sender
    em['to'] = email_receiver
    em['subject'] = subject

    # Add text/plain part (the message body)
    em.attach(MIMEText(body, 'plain'))

    # Attach a file (e.g., attachment.txt)
    attachment_path = 'apartment_orders.xlsx'
    attachment_name = 'apartment_orders.xlsx'

    with open(attachment_path, 'rb') as attachment_file:
        # Create a base MIME type for the attachment
        attachment = MIMEBase('application', 'octet-stream')
        attachment.set_payload(attachment_file.read())

    # Encode the attachment in ASCII
    encoders.encode_base64(attachment)

    # Set the filename and add the attachment to the message
    attachment.add_header('Content-Disposition', f'attachment; filename={attachment_name}')
    em.attach(attachment)

    context = ssl.create_default_context()

    with smtplib.SMTP_SSL('smtp.gmail.com', 465, context=context) as smtp:
        smtp.login(email_sender, email_password)
        smtp.sendmail(email_sender, email_receiver, em.as_string())


def createExcel():
    # Read the data.json file to get product names and prices
    with open('data.json', 'r', encoding='utf-8') as data_file:
        products = json.load(data_file)

    # Create a dictionary to store product prices
    product_prices = {product['name']: product['price'] for product in products}

    # Create a list to store apartment data
    apartment_data = []

    # Define the directory containing apartment JSON files
    directory = 'checkout/orders/'

    # Iterate through JSON files in the directory
    for filename in os.listdir(directory):
        if filename.endswith('.json'):
            apartment_number = os.path.splitext(filename)[0]
            with open(os.path.join(directory, filename), 'r', encoding='utf-8') as apartment_file:
                apartment_order = json.load(apartment_file)
                second_name = apartment_order['name']
                order_items = apartment_order['order']

                # Create a dictionary to store the apartment's order
                apartment_order_data = {'Appartment': apartment_number, 'Nachname': second_name}

                # Add product amounts to the dictionary
                for item in order_items:
                    product_id = item['id']
                    amount = item['amount']
                    product_name = next((p['name'] for p in products if p['id'] == product_id), '')

                    apartment_order_data[product_name] = amount

                total_price = 0
                for item in order_items:
                    product_id = item['id']
                    amount = item['amount']
                    product_name = next((p['name'] for p in products if p['id'] == product_id), '')
                    product_price = product_prices.get(product_name, 0)
                    total_price += amount * product_price

                apartment_order_data['Preis'] = f"{total_price:0.2F}€".replace(",",
                                                                               ".")  # Add the total price to the dictionary

                apartment_data.append(apartment_order_data)

    # Create a DataFrame from the apartment data
    df = pd.DataFrame(apartment_data)

    # Reorder the columns to have 'Preis' as the last column
    column_order = [col for col in df.columns if col != 'Preis'] + ['Preis']
    df = df[column_order]

    # Save the DataFrame to an Excel file
    excel_file = 'apartment_orders.xlsx'
    df.to_excel(excel_file, index=False, engine='openpyxl')

    # Load the workbook
    workbook = load_workbook(excel_file)
    worksheet = workbook.active

    # Adjust column widths to fit the content
    for column in worksheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2  # Adjust this value as needed
        worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    # Save the modified workbook
    workbook.save(excel_file)


createExcel()
sendMail()
